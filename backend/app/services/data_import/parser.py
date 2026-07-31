from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from math import isfinite
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from app.services.data_import.templates import (
    ColumnDefinition,
    TemplateMatch,
    detect_template,
    normalize_header_value,
)

MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_DATA_ROWS = 10_000
MAX_COLUMNS = 100
SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}
MISSING_MARKERS = {"", "-", "--", "N/A", "n/a", "null", "None"}
TOO_MANY_COLUMNS_MESSAGE = "Files with more than 100 columns are not supported"
TOO_MANY_ROWS_MESSAGE = "Files with more than 10,000 rows are not supported"
MAX_ARCHIVE_ENTRY_COUNT = 2_048
MAX_ARCHIVE_TOTAL_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ARCHIVE_ENTRY_UNCOMPRESSED_BYTES = 20 * 1024 * 1024
MAX_ARCHIVE_COMPRESSION_RATIO = 100
FORMULA_ELEMENT_PATTERN = re.compile(rb"<(?:[A-Za-z_][\w.-]*:)?f(?=[\s/>])")
GROUPED_NUMBER_PATTERN = re.compile(r"^[+-]?\d{1,3}(?:,\d{3})+(?:\.\d+)?$")
SQL_INTEGER_MIN = -(2**31)
SQL_INTEGER_MAX = 2**31 - 1


class ParseFailure(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


@dataclass(frozen=True, slots=True)
class RowIssue:
    code: str
    message: str
    field: str | None = None


@dataclass(slots=True)
class NormalizedRow:
    template_code: str
    row_number: int
    raw: dict[str, Any]
    normalized: dict[str, Any]
    errors: list[RowIssue] = field(default_factory=list)
    warnings: list[RowIssue] = field(default_factory=list)


ValidatedRow = NormalizedRow


@dataclass(frozen=True, slots=True)
class ImportPreview:
    template_code: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    warning_rows: int


@dataclass(frozen=True, slots=True)
class ParsedDataset:
    template_code: str
    template_name: str
    headers: list[str]
    rows: list[ValidatedRow]
    preview: ImportPreview
    sheet_name: str | None = None
    dataset_ordinal: int = 1
    warnings: list[RowIssue] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class ParsedDatasetFailure:
    sheet_name: str | None
    dataset_ordinal: int
    code: str
    message: str


@dataclass(frozen=True, slots=True)
class ParsedSourceFile:
    filename: str
    datasets: list[ParsedDataset]
    warnings: list[RowIssue] = field(default_factory=list)
    failures: list[ParsedDatasetFailure] = field(default_factory=list)

    def require_single_dataset(self) -> ParsedDataset:
        if self.failures:
            raise ParseFailure(
                "partial_workbook_requires_bulk_import",
                "A workbook with failed worksheets must use the bulk import flow",
            )
        if len(self.datasets) != 1:
            raise ParseFailure(
                "multiple_datasets_require_bulk_import",
                "Files containing multiple datasets must use the bulk import flow",
            )
        return self.datasets[0]

    @property
    def template_code(self) -> str:
        return self.require_single_dataset().template_code

    @property
    def template_name(self) -> str:
        return self.require_single_dataset().template_name

    @property
    def headers(self) -> list[str]:
        return self.require_single_dataset().headers

    @property
    def rows(self) -> list[ValidatedRow]:
        return self.require_single_dataset().rows

    @property
    def preview(self) -> ImportPreview:
        return self.require_single_dataset().preview


@dataclass(frozen=True, slots=True)
class _SourceTable:
    sheet_name: str | None
    dataset_ordinal: int
    headers: list[str]
    raw_rows: list[dict[str, Any]]


def parse_source_file(filename: str, data: bytes) -> ParsedSourceFile:
    extension = _validate_source(filename, data)
    source_warnings: list[RowIssue] = []
    failures: list[ParsedDatasetFailure] = []
    if extension == ".xlsx":
        tables, source_warnings, failures = _parse_xlsx(data)
    else:
        headers, raw_rows = _parse_csv(data)
        tables = [
            _SourceTable(
                sheet_name=None,
                dataset_ordinal=1,
                headers=headers,
                raw_rows=raw_rows,
            )
        ]

    datasets: list[ParsedDataset] = []
    for table in tables:
        try:
            match = _detect_template_or_fail(table.headers)
        except ParseFailure as exc:
            failures.append(
                ParsedDatasetFailure(
                    sheet_name=table.sheet_name,
                    dataset_ordinal=table.dataset_ordinal,
                    code=exc.code,
                    message=str(exc),
                )
            )
            continue
        template = match.template
        rows = _normalize_rows(match, table.headers, table.raw_rows)
        preview = build_preview(rows, template_code=template.code)
        warnings = [
            RowIssue(
                code="ignored_column",
                message=f"Unrecognized column was preserved for audit: {header}",
                field=header,
            )
            for header in match.ignored_headers
        ]
        datasets.append(
            ParsedDataset(
                template_code=template.code,
                template_name=template.display_name,
                headers=table.headers,
                rows=rows,
                preview=preview,
                sheet_name=table.sheet_name,
                dataset_ordinal=table.dataset_ordinal,
                warnings=warnings,
            )
        )

    if not datasets:
        if failures:
            first = failures[0]
            raise ParseFailure(first.code, first.message)
        raise ParseFailure("empty_file", "The file does not contain any rows")
    return ParsedSourceFile(
        filename=filename,
        datasets=datasets,
        warnings=source_warnings,
        failures=failures,
    )


def parse_single_source_file(filename: str, data: bytes) -> ParsedDataset:
    return parse_source_file(filename, data).require_single_dataset()


def build_preview(
    rows: list[ValidatedRow],
    *,
    template_code: str | None = None,
) -> ImportPreview:
    resolved_template_code = template_code or (rows[0].template_code if rows else "")
    invalid_rows = sum(1 for row in rows if row.errors)
    warning_rows = sum(1 for row in rows if row.warnings)
    return ImportPreview(
        template_code=resolved_template_code,
        total_rows=len(rows),
        valid_rows=len(rows) - invalid_rows,
        invalid_rows=invalid_rows,
        warning_rows=warning_rows,
    )


def _validate_source(filename: str, data: bytes) -> str:
    if "\x00" in filename:
        raise ParseFailure(
            "invalid_filename",
            "Embedded NUL bytes are not allowed in filenames",
        )
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ParseFailure("unsupported_extension", "Unsupported file extension")
    if len(data) > MAX_FILE_BYTES:
        raise ParseFailure("file_too_large", "Files larger than 10 MB are not supported")
    return extension


def _parse_csv(data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = _decode_csv(data)
    stream = io.StringIO(text, newline="")
    try:
        reader = csv.reader(stream, strict=True)
        rows = list(reader)
    except csv.Error as exc:
        raise ParseFailure("malformed_csv", f"Malformed CSV: {exc}") from exc
    return _rows_to_table(rows, strict_row_width=True)


def _decode_csv(data: bytes) -> str:
    try:
        if data.startswith(b"\xef\xbb\xbf"):
            text = data.decode("utf-8-sig")
        else:
            text = data.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ParseFailure("invalid_encoding", "CSV files must use UTF-8 or UTF-8 BOM") from exc
    if "\x00" in text:
        raise ParseFailure("embedded_nul", "Embedded NUL bytes are not allowed")
    return text


def _parse_xlsx(
    data: bytes,
) -> tuple[list[_SourceTable], list[RowIssue], list[ParsedDatasetFailure]]:
    _scan_xlsx_archive(data)
    try:
        workbook = load_workbook(
            filename=io.BytesIO(data),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except BadZipFile as exc:
        raise ParseFailure("invalid_xlsx", "The workbook is not a valid XLSX file") from exc
    except Exception as exc:  # pragma: no cover
        raise ParseFailure("invalid_xlsx", f"Unable to read workbook: {exc}") from exc

    try:
        if not workbook.worksheets:
            raise ParseFailure("empty_workbook", "The workbook does not contain any worksheets")
        tables: list[_SourceTable] = []
        warnings: list[RowIssue] = []
        failures: list[ParsedDatasetFailure] = []
        used_sheet_names: set[str] = set()
        total_data_rows = 0
        for dataset_ordinal, worksheet in enumerate(workbook.worksheets, start=1):
            sheet_name = _stable_sheet_name(worksheet.title, used_sheet_names)
            rows: list[list[Any]] = []
            for row in worksheet.iter_rows(values_only=True):
                values = list(row)
                last_index = max(
                    (
                        index
                        for index, value in enumerate(values)
                        if value is not None and str(value).strip() != ""
                    ),
                    default=-1,
                )
                if last_index < 0:
                    continue
                trimmed = values[: last_index + 1]
                if len(trimmed) > MAX_COLUMNS:
                    failures.append(
                        ParsedDatasetFailure(
                            sheet_name=sheet_name,
                            dataset_ordinal=dataset_ordinal,
                            code="too_many_columns",
                            message=TOO_MANY_COLUMNS_MESSAGE,
                        )
                    )
                    rows = []
                    break
                rows.append(trimmed)
                if len(rows) - 1 > MAX_DATA_ROWS:
                    raise ParseFailure("too_many_rows", TOO_MANY_ROWS_MESSAGE)
            if not rows:
                if any(
                    failure.dataset_ordinal == dataset_ordinal
                    for failure in failures
                ):
                    continue
                warnings.append(
                    RowIssue(
                        code="blank_worksheet_skipped",
                        message=f"Blank worksheet was skipped: {sheet_name}",
                        field=sheet_name,
                    )
                )
                continue
            try:
                headers, raw_rows = _rows_to_table(rows, reject_extra_cells=True)
            except ParseFailure as exc:
                failures.append(
                    ParsedDatasetFailure(
                        sheet_name=sheet_name,
                        dataset_ordinal=dataset_ordinal,
                        code=exc.code,
                        message=str(exc),
                    )
                )
                continue
            total_data_rows += len(raw_rows)
            if total_data_rows > MAX_DATA_ROWS:
                raise ParseFailure("too_many_rows", TOO_MANY_ROWS_MESSAGE)
            tables.append(
                _SourceTable(
                    sheet_name=sheet_name,
                    dataset_ordinal=dataset_ordinal,
                    headers=headers,
                    raw_rows=raw_rows,
                )
            )
        return tables, warnings, failures
    finally:
        workbook.close()


def _stable_sheet_name(raw_name: str, used_names: set[str]) -> str:
    base_name = raw_name.strip() or "Sheet"
    candidate = base_name
    suffix = 2
    while candidate.casefold() in used_names:
        candidate = f"{base_name} ({suffix})"
        suffix += 1
    used_names.add(candidate.casefold())
    return candidate


def _scan_xlsx_archive(data: bytes) -> None:
    try:
        with ZipFile(io.BytesIO(data)) as archive:
            infos = archive.infolist()
            _validate_archive_metadata(infos)
            names = [info.filename for info in infos]
            if any(name.startswith("xl/externalLinks/") for name in names):
                raise ParseFailure(
                    "external_links_unsupported",
                    "Workbooks with external links are not supported",
                )
            if any(name.endswith("vbaProject.bin") for name in names):
                raise ParseFailure(
                    "macros_unsupported",
                    "Macro-enabled workbooks are not supported",
                )
            for info in infos:
                if not info.filename.startswith("xl/worksheets/") or not info.filename.endswith(
                    ".xml"
                ):
                    continue
                if _entry_contains_formula(archive, info):
                    raise ParseFailure(
                        "formula_cells_unsupported",
                        "Formula cells are not supported",
                    )
    except BadZipFile as exc:
        raise ParseFailure("invalid_xlsx", "The workbook is not a valid XLSX file") from exc


def _rows_to_table(
    rows: list[list[Any]],
    *,
    strict_row_width: bool = False,
    reject_extra_cells: bool = False,
) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        raise ParseFailure("empty_file", "The file does not contain any rows")
    headers = [_stringify_header(value) for value in rows[0]]
    _validate_headers(headers)
    data_rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows[1:], start=2):
        if strict_row_width and len(values) != len(headers):
            raise ParseFailure(
                "csv_row_width_mismatch",
                f"CSV row {row_number} has {len(values)} fields; expected {len(headers)} fields",
            )
        if reject_extra_cells and len(values) > len(headers):
            raise ParseFailure(
                "xlsx_row_width_mismatch",
                (
                    f"XLSX row {row_number} has {len(values)} fields; "
                    f"expected at most {len(headers)}"
                ),
            )
        trimmed = list(values if strict_row_width else values[: len(headers)])
        if not strict_row_width and len(trimmed) < len(headers):
            trimmed.extend([None] * (len(headers) - len(trimmed)))
        if _row_is_blank(trimmed):
            continue
        data_rows.append({"row_number": row_number, "values": trimmed})
        if len(data_rows) > MAX_DATA_ROWS:
            raise ParseFailure("too_many_rows", TOO_MANY_ROWS_MESSAGE)
    return headers, data_rows


def _stringify_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("\ufeff", "").strip()


def _validate_headers(headers: list[str]) -> None:
    if len(headers) > MAX_COLUMNS:
        raise ParseFailure("too_many_columns", TOO_MANY_COLUMNS_MESSAGE)
    normalized = [normalize_header_value(header) for header in headers]
    seen: set[str] = set()
    duplicates: set[str] = set()
    for header in normalized:
        if not header:
            continue
        if header in seen:
            duplicates.add(header)
        seen.add(header)
    if duplicates:
        raise ParseFailure("duplicate_headers", "Duplicate headers are not allowed")
    if any(not header for header in normalized):
        raise ParseFailure("blank_headers", "Blank headers are not supported")


def _detect_template_or_fail(headers: list[str]) -> TemplateMatch:
    try:
        return detect_template(headers)
    except ValueError as exc:
        if str(exc) == "ambiguous":
            raise ParseFailure("ambiguous_template", "Ambiguous template signature") from exc
        if str(exc) == "duplicate_canonical_field":
            raise ParseFailure(
                "duplicate_canonical_field",
                "Multiple columns map to the same field",
            ) from exc
        raise ParseFailure("unknown_template", "Unknown or unsupported template") from exc


def _normalize_rows(
    match: TemplateMatch,
    headers: list[str],
    raw_rows: list[dict[str, Any]],
) -> list[ValidatedRow]:
    template = match.template
    ignored_warnings = [
        RowIssue(
            code="ignored_column",
            message=f"Unrecognized column was preserved for audit: {header}",
            field=header,
        )
        for header in match.ignored_headers
    ]
    rows: list[ValidatedRow] = []
    for row in raw_rows:
        row_values = row["values"]
        raw = {
            header: row_values[index] if index < len(row_values) else None
            for index, header in enumerate(headers)
        }
        normalized: dict[str, Any] = {}
        errors: list[RowIssue] = []
        warnings = list(ignored_warnings)
        for column in template.columns:
            source_index = match.column_indexes.get(column.field_name)
            raw_value = (
                row_values[source_index]
                if source_index is not None and source_index < len(row_values)
                else None
            )
            parsed_value, issues = _parse_column_value(column, raw_value)
            normalized.update(parsed_value)
            errors.extend(issues)
        rows.append(
            ValidatedRow(
                template_code=template.code,
                row_number=row["row_number"],
                raw=raw,
                normalized=normalized,
                errors=errors,
                warnings=warnings,
            )
        )
    return rows


def _parse_column_value(
    column: ColumnDefinition,
    raw_value: Any,
) -> tuple[dict[str, Any], list[RowIssue]]:
    field_name = column.field_name
    if column.value_type == "string":
        normalized_string = _normalize_string(raw_value)
        parsed_value, issues = {field_name: normalized_string}, []
        if (
            normalized_string is not None
            and column.max_length is not None
            and len(normalized_string) > column.max_length
        ):
            issues.append(_issue("value_too_long", field_name))
    elif column.value_type == "int":
        parsed_value, issues = _parse_int(field_name, raw_value, minimum=column.minimum)
    elif column.value_type == "float":
        parsed_value, issues = _parse_float(field_name, raw_value, minimum=column.minimum)
    elif column.value_type == "ratio":
        parsed_value, issues = _parse_ratio(field_name, raw_value)
    elif column.value_type == "date":
        parsed_value, issues = _parse_date(field_name, raw_value)
    elif column.value_type == "datetime":
        parsed_value, issues = _parse_datetime(field_name, raw_value)
    elif column.value_type == "date_range":
        parsed_value, issues = _parse_date_range(raw_value)
    else:
        raise ParseFailure(
            "unsupported_column_type",
            f"Unsupported column type: {column.value_type}",
        )
    if column.required and _coerce_text(raw_value) is None:
        issues.append(_issue("required_value_missing", field_name))
    return parsed_value, issues


def _normalize_string(value: Any) -> str | None:
    text = _coerce_text(value)
    return None if text is None else text.strip()


def _parse_int(
    field_name: str,
    value: Any,
    *,
    minimum: float | None = None,
) -> tuple[dict[str, Any], list[RowIssue]]:
    text = _coerce_text(value)
    if text is None:
        return {field_name: None}, []
    try:
        number = _parse_number_text(text)
    except ValueError:
        return {field_name: None}, [_issue("invalid_integer", field_name)]
    if not isfinite(number) or not number.is_integer():
        return {field_name: None}, [_issue("invalid_integer", field_name)]
    if number < SQL_INTEGER_MIN or number > SQL_INTEGER_MAX:
        return {field_name: None}, [_issue("integer_out_of_range", field_name)]
    if minimum is not None and number < minimum:
        return {field_name: None}, [_issue("below_minimum", field_name)]
    return {field_name: int(number)}, []


def _parse_float(
    field_name: str,
    value: Any,
    *,
    minimum: float | None = None,
) -> tuple[dict[str, Any], list[RowIssue]]:
    text = _coerce_text(value)
    if text is None:
        return {field_name: None}, []
    try:
        number = _parse_number_text(text)
    except ValueError:
        return {field_name: None}, [_issue("invalid_number", field_name)]
    if not isfinite(number):
        return {field_name: None}, [_issue("invalid_number", field_name)]
    if minimum is not None and number < minimum:
        return {field_name: None}, [_issue("below_minimum", field_name)]
    return {field_name: number}, []


def _parse_ratio(field_name: str, value: Any) -> tuple[dict[str, Any], list[RowIssue]]:
    text = _coerce_text(value)
    if text is None:
        return {field_name: None}, []
    try:
        number = (
            _parse_number_text(text[:-1]) / 100
            if text.endswith("%")
            else _parse_number_text(text)
        )
    except ValueError:
        return {field_name: None}, [_issue("invalid_number", field_name)]
    if not isfinite(number):
        return {field_name: None}, [_issue("invalid_number", field_name)]
    if not 0 <= number <= 1:
        return {field_name: None}, [_issue("out_of_range", field_name)]
    return {field_name: number}, []


def _parse_date(field_name: str, value: Any) -> tuple[dict[str, Any], list[RowIssue]]:
    if isinstance(value, datetime):
        return {field_name: value.date()}, []
    if isinstance(value, date):
        return {field_name: value}, []
    text = _coerce_text(value)
    if text is None:
        return {field_name: None}, []
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return {field_name: datetime.strptime(text, fmt).date()}, []
        except ValueError:
            continue
    return {field_name: None}, [_issue("invalid_date", field_name)]


def _parse_datetime(field_name: str, value: Any) -> tuple[dict[str, Any], list[RowIssue]]:
    if isinstance(value, datetime):
        return {field_name: value}, []
    if isinstance(value, date):
        return {field_name: datetime.combine(value, datetime.min.time())}, []
    text = _coerce_text(value)
    if text is None:
        return {field_name: None}, []
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return {field_name: datetime.strptime(text, fmt)}, []
        except ValueError:
            continue
    return {field_name: None}, [_issue("invalid_datetime", field_name)]


def _parse_date_range(value: Any) -> tuple[dict[str, Any], list[RowIssue]]:
    text = _coerce_text(value)
    if text is None:
        return {"period_start": None, "period_end": None}, []
    separator = " ~ " if " ~ " in text else "~"
    parts = [part.strip() for part in text.split(separator)]
    if len(parts) != 2:
        return (
            {"period_start": None, "period_end": None},
            [RowIssue(code="invalid_date_range", message="Invalid date range", field="period")],
        )
    start_value, start_errors = _parse_date("period_start", parts[0])
    end_value, end_errors = _parse_date("period_end", parts[1])
    errors = [*start_errors, *end_errors]
    start = start_value["period_start"]
    end = end_value["period_end"]
    if not errors and start is not None and end is not None and start > end:
        errors.append(
            RowIssue(
                code="reversed_date_range",
                message="Date range start must not be later than its end",
                field="period",
            )
        )
    return {**start_value, **end_value}, errors


def _coerce_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if text in MISSING_MARKERS:
        return None
    return text


def _parse_number_text(text: str) -> float:
    if "," in text:
        if not GROUPED_NUMBER_PATTERN.fullmatch(text):
            raise ValueError("invalid thousands separators")
        text = text.replace(",", "")
    return float(text)


def _row_is_blank(values: list[Any]) -> bool:
    return all(_coerce_text(value) is None for value in values)


def _issue(code: str, field_name: str) -> RowIssue:
    messages = {
        "invalid_integer": f"Invalid integer for {field_name}",
        "invalid_number": f"Invalid number for {field_name}",
        "out_of_range": f"{field_name} must be between 0 and 1",
        "below_minimum": f"{field_name} must not be negative",
        "integer_out_of_range": f"{field_name} is outside the supported integer range",
        "required_value_missing": f"{field_name} is required",
        "value_too_long": f"{field_name} is too long",
        "invalid_date": f"Invalid date for {field_name}",
        "invalid_datetime": f"Invalid datetime for {field_name}",
    }
    return RowIssue(code=code, message=messages[code], field=field_name)


def _validate_archive_metadata(infos: list[Any]) -> None:
    if len(infos) > MAX_ARCHIVE_ENTRY_COUNT:
        raise ParseFailure(
            "archive_too_many_entries",
            f"XLSX archive has too many archive entries ({len(infos)} > {MAX_ARCHIVE_ENTRY_COUNT})",
        )

    total_uncompressed_bytes = 0
    seen_filenames: set[str] = set()
    for info in infos:
        if info.is_dir():
            continue
        if info.filename in seen_filenames:
            raise ParseFailure(
                "archive_duplicate_entries",
                f"XLSX archive contains duplicate archive entries for {info.filename}",
            )
        seen_filenames.add(info.filename)
        if info.file_size > MAX_ARCHIVE_ENTRY_UNCOMPRESSED_BYTES:
            raise ParseFailure(
                "archive_entry_too_large",
                "XLSX archive entry is too large before decompression",
            )
        total_uncompressed_bytes += info.file_size
        if total_uncompressed_bytes > MAX_ARCHIVE_TOTAL_UNCOMPRESSED_BYTES:
            raise ParseFailure(
                "archive_total_uncompressed_too_large",
                "XLSX archive total uncompressed size exceeds the allowed limit",
            )
        if _compression_ratio(info) > MAX_ARCHIVE_COMPRESSION_RATIO:
            raise ParseFailure(
                "archive_compression_ratio_too_large",
                "XLSX archive entry exceeds the allowed compression ratio",
            )


def _compression_ratio(info: Any) -> float:
    if info.file_size == 0:
        return 0.0
    if info.compress_size == 0:
        return float("inf")
    return info.file_size / info.compress_size


def _entry_contains_formula(archive: ZipFile, info: Any) -> bool:
    tail = b""
    with archive.open(info, "r") as entry:
        while True:
            chunk = entry.read(8_192)
            if not chunk:
                return False
            haystack = tail + chunk
            if FORMULA_ELEMENT_PATTERN.search(haystack):
                return True
            tail = haystack[-64:]
