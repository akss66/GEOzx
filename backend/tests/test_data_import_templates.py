from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile

import pytest
from openpyxl import Workbook

import app.services.data_import.parser as parser_module
from app.services.data_import.adapters import FileDataSourceAdapter
from app.services.data_import.parser import ParseFailure, parse_source_file

DAILY_HEADERS = ["日期", "播放量"]
SINGLE_CONTENT_HEADERS = [
    "视频名称",
    "发布时间",
    "播放量",
    "5s完播率",
    "2s跳出率",
    "平均播放时长",
]
PERIOD_AGGREGATE_HEADERS = [
    "发布时间",
    "体裁",
    "垂类",
    "周期内投稿量",
    "条均点击率",
    "条均5s完播率",
    "条均2s跳出率",
    "条均播放时长",
    "播放量中位数",
    "条均点赞数",
    "条均评论量",
    "条均分享量",
]
WORK_LIST_HEADERS = [
    "作品名称",
    "发布时间",
    "体裁",
    "审核状态",
    "播放量",
    "完播率",
    "5s完播率",
    "封面点击率",
    "2s跳出率",
    "平均播放时长",
    "点赞量",
    "分享量",
    "评论量",
    "收藏量",
    "主页访问量",
    "粉丝增量",
]

DOUYIN_DAILY_ACCOUNT_EXPORT_CASES = [
    ("主页访问", "profile_visit_count", 12, 12),
    ("总粉丝量", "follower_count", 7994, 7994),
    ("取关粉丝", "unfollow_count", 3, 3),
    ("净增粉丝", "follower_delta", -11, -11),
    ("封面点击率", "cover_click_rate", "83.33%", pytest.approx(0.8333)),
    ("作品评论", "comment_count", -2, -2),
    ("作品分享", "share_count", 5, 5),
    ("作品点赞", "like_count", -21, -21),
]

CONTENT_TYPES_XML = """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels"
   ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml"
   ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml"
   ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>
"""

ROOT_RELS_XML = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
   Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
   Target="xl/workbook.xml"/>
</Relationships>
"""

WORKBOOK_XML = """<?xml version="1.0" encoding="UTF-8"?>
<workbook
 xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>
"""

WORKBOOK_RELS_XML = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
   Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
   Target="worksheets/sheet1.xml"/>
</Relationships>
"""


def workbook_bytes(
    headers: list[str],
    rows: Iterable[Iterable[object | None]],
    *,
    formulas: set[tuple[int, int]] | None = None,
    extras: dict[str, bytes | tuple[bytes, int]] | None = None,
) -> bytes:
    formulas = formulas or set()
    buffer = io.BytesIO()
    sheet_rows = [_xml_row(1, headers, formulas)]
    for index, row in enumerate(rows, start=2):
        sheet_rows.append(_xml_row(index, list(row), formulas))
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData>"
        f"{''.join(sheet_rows)}"
        "</sheetData>"
        "</worksheet>"
    )

    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", CONTENT_TYPES_XML)
        archive.writestr("_rels/.rels", ROOT_RELS_XML)
        archive.writestr("xl/workbook.xml", WORKBOOK_XML)
        archive.writestr("xl/_rels/workbook.xml.rels", WORKBOOK_RELS_XML)
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        for filename, content in (extras or {}).items():
            if isinstance(content, tuple):
                body, compress_type = content
                archive.writestr(filename, body, compress_type=compress_type)
            else:
                archive.writestr(filename, content)

    return buffer.getvalue()


@pytest.mark.parametrize(
    ("metric_header", "field_name", "raw_value", "expected"),
    DOUYIN_DAILY_ACCOUNT_EXPORT_CASES,
)
def test_real_douyin_daily_account_metric_exports_are_recognized(
    metric_header, field_name, raw_value, expected
):
    parsed = parse_source_file(
        f"数据表现_{metric_header}数据.xlsx",
        workbook_bytes(["日期", metric_header], [["2026-07-30", raw_value]]),
    ).require_single_dataset()

    assert parsed.template_code.startswith("douyin_daily_")
    assert parsed.preview.valid_rows == 1
    assert parsed.rows[0].normalized["stat_date"] == date(2026, 7, 30)
    assert parsed.rows[0].normalized[field_name] == expected


def csv_bytes(
    headers: list[str],
    rows: Iterable[Iterable[object | None]],
    *,
    bom: bool = False,
) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    data = buffer.getvalue().encode()
    return (b"\xef\xbb\xbf" + data) if bom else data


def typed_workbook_bytes(
    headers: list[str],
    rows: Iterable[Iterable[object | None]],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(list(row))
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _xml_row(index: int, values: list[object | None], formulas: set[tuple[int, int]]) -> str:
    cells: list[str] = []
    for column_index, value in enumerate(values, start=1):
        ref = f"{_column_label(column_index)}{index}"
        if (index, column_index) in formulas:
            cells.append(f'<c r="{ref}"><f>SUM(1,2)</f><v>3</v></c>')
            continue
        if value is None:
            cells.append(f'<c r="{ref}"/>')
            continue
        text = str(value)
        if _looks_numeric(text):
            cells.append(f'<c r="{ref}"><v>{text}</v></c>')
            continue
        escaped = (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
        )
        cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{escaped}</t></is></c>')
    return f'<row r="{index}">{"".join(cells)}</row>'


def _looks_numeric(value: str) -> bool:
    if value == "":
        return False
    try:
        float(value)
    except ValueError:
        return False
    return True


def _column_label(index: int) -> str:
    result = ""
    current = index
    while current:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _fail_if_opened(*_args, **_kwargs):
    raise AssertionError("archive entry should not be decompressed")


def _fail_if_loaded(*_args, **_kwargs):
    raise AssertionError("openpyxl should not run for this rejection path")


@pytest.mark.parametrize(
    ("filename", "payload"),
    [
        ("daily.xlsx", lambda: workbook_bytes(DAILY_HEADERS, [["2026-07-18", "81"]])),
        ("daily.csv", lambda: csv_bytes(DAILY_HEADERS, [["2026-07-18", "81"]])),
        (
            "daily-bom.csv",
            lambda: csv_bytes(DAILY_HEADERS, [["2026-07-18", "81"]], bom=True),
        ),
    ],
)
def test_daily_play_template_detects_in_excel_and_csv(filename: str, payload):
    parsed = parse_source_file(filename, payload())

    assert parsed.template_code == "douyin_daily_play_v1"
    assert parsed.preview.total_rows == 1
    assert parsed.rows[0].normalized["stat_date"] == date(2026, 7, 18)
    assert parsed.rows[0].normalized["play"] == 81
    assert parsed.rows[0].errors == []


def test_daily_play_accepts_typed_excel_dates_and_grouped_counts():
    typed = parse_source_file(
        "daily.xlsx",
        typed_workbook_bytes(DAILY_HEADERS, [[date(2026, 7, 18), 1234]]),
    )
    grouped = parse_source_file(
        "daily.csv",
        csv_bytes(DAILY_HEADERS, [["2026-07-18", "1,234"]]),
    )

    assert typed.rows[0].normalized == {"stat_date": date(2026, 7, 18), "play": 1234}
    assert grouped.rows[0].normalized["play"] == 1234


@pytest.mark.parametrize(
    ("filename", "payload"),
    [
        (
            "single.xlsx",
            lambda: workbook_bytes(
                SINGLE_CONTENT_HEADERS,
                [["燃尽。#codex", "2026-07-07 13:36", "444", "0.1471", "0.6353", "2.8314"]],
            ),
        ),
        (
            "single.csv",
            lambda: csv_bytes(
                SINGLE_CONTENT_HEADERS,
                [["燃尽。#codex", "2026-07-07 13:36", "444", "0.1471", "0.6353", "2.8314"]],
            ),
        ),
    ],
)
def test_single_content_template_normalizes_metrics(filename: str, payload):
    parsed = parse_source_file(filename, payload())

    assert parsed.template_code == "douyin_single_content_v1"
    assert parsed.rows[0].normalized == {
        "title": "燃尽。#codex",
        "published_at": datetime(2026, 7, 7, 13, 36),
        "play": 444,
        "completion_rate_5s": 0.1471,
        "bounce_rate_2s": 0.6353,
        "avg_watch_time_seconds": 2.8314,
    }


@pytest.mark.parametrize(
    ("filename", "payload"),
    [
        (
            "aggregate.xlsx",
            lambda: workbook_bytes(
                PERIOD_AGGREGATE_HEADERS,
                [[
                    "2026-04-23 ~ 2026-07-22",
                    "1min-视频,图文",
                    "随拍",
                    "1",
                    "0.0000",
                    "0.1471",
                    "0.6353",
                    "2.8314",
                    "444.0000",
                    "10.0000",
                    "0.0000",
                    "1.0000",
                ]],
            ),
        ),
        (
            "aggregate.csv",
            lambda: csv_bytes(
                PERIOD_AGGREGATE_HEADERS,
                [[
                    "2026-04-23 ~ 2026-07-22",
                    "1min-视频,图文",
                    "随拍",
                    "1",
                    "0.0000",
                    "0.1471",
                    "0.6353",
                    "2.8314",
                    "444.0000",
                    "10.0000",
                    "0.0000",
                    "1.0000",
                ]],
            ),
        ),
    ],
)
def test_period_aggregate_template_parses_date_range_and_numeric_fields(
    filename: str,
    payload,
):
    parsed = parse_source_file(filename, payload())

    assert parsed.template_code == "douyin_period_aggregate_v1"
    assert parsed.rows[0].normalized["period_start"] == date(2026, 4, 23)
    assert parsed.rows[0].normalized["period_end"] == date(2026, 7, 22)
    assert parsed.rows[0].normalized["publish_count"] == 1
    assert parsed.rows[0].normalized["median_play"] == 444
    assert parsed.rows[0].normalized["avg_share_count"] == 1


def test_period_aggregate_accepts_fractional_median_and_average_counts():
    parsed = parse_source_file(
        "aggregate.xlsx",
        workbook_bytes(
            PERIOD_AGGREGATE_HEADERS,
            [
                [
                    "2026-05-02 ~ 2026-07-31",
                    "1min-视频,图文",
                    "科技,随拍",
                    "2",
                    "0.1354",
                    "0.1630",
                    "0.6135",
                    "3.2148",
                    "398.5000",
                    "8.2500",
                    "1.5000",
                    "0.5000",
                ]
            ],
        ),
    )

    row = parsed.rows[0]
    assert row.errors == []
    assert row.normalized["median_play"] == 398.5
    assert row.normalized["avg_like_count"] == 8.25
    assert row.normalized["avg_comment_count"] == 1.5
    assert row.normalized["avg_share_count"] == 0.5


def test_period_aggregate_rejects_reversed_date_ranges():
    parsed = parse_source_file(
        "aggregate.xlsx",
        workbook_bytes(
            PERIOD_AGGREGATE_HEADERS,
            [[
                "2026-07-31 ~ 2026-05-02",
                "1min-视频",
                "科技",
                "2",
                "0.1354",
                "0.1630",
                "0.6135",
                "3.2148",
                "398.5000",
                "8.0000",
                "1.5000",
                "0.5000",
            ]],
        ),
    )

    assert parsed.preview.invalid_rows == 1
    assert {issue.code for issue in parsed.rows[0].errors} == {"reversed_date_range"}


@pytest.mark.parametrize(
    ("filename", "payload"),
    [
        (
            "works.xlsx",
            lambda: workbook_bytes(
                WORK_LIST_HEADERS,
                [[
                    "作品 A",
                    "2026-07-18 14:11:20",
                    "1min-视频",
                    "公开",
                    "81",
                    "0.087500",
                    "0.375000",
                    "-",
                    "0.375000",
                    "9.53",
                    "6",
                    "0",
                    "3",
                    "0",
                    "3",
                    "0",
                ]],
            ),
        ),
        (
            "works.csv",
            lambda: csv_bytes(
                WORK_LIST_HEADERS,
                [[
                    "作品 A",
                    "2026-07-18 14:11:20",
                    "1min-视频",
                    "公开",
                    "81",
                    "0.087500",
                    "0.375000",
                    "-",
                    "0.375000",
                    "9.53",
                    "6",
                    "0",
                    "3",
                    "0",
                    "3",
                    "0",
                ]],
            ),
        ),
    ],
)
def test_work_list_template_normalizes_percentages_and_missing_values(
    filename: str,
    payload,
):
    parsed = parse_source_file(filename, payload())

    assert parsed.template_code == "douyin_work_list_v1"
    assert parsed.rows[0].normalized["title"] == "作品 A"
    assert parsed.rows[0].normalized["published_at"] == datetime(2026, 7, 18, 14, 11, 20)
    assert parsed.rows[0].normalized["play"] == 81
    assert parsed.rows[0].normalized["completion_rate"] == 0.0875
    assert parsed.rows[0].normalized["completion_rate_5s"] == 0.375
    assert parsed.rows[0].normalized["cover_click_rate"] is None
    assert parsed.rows[0].normalized["profile_visit_count"] == 3
    assert parsed.rows[0].normalized["follower_delta"] == 0


def test_invalid_dates_and_percentages_are_reported_in_preview():
    parsed = parse_source_file(
        "works.csv",
        csv_bytes(
            WORK_LIST_HEADERS,
            [[
                "作品 A",
                "not-a-date",
                "1min-视频",
                "公开",
                "81",
                "1.25",
                "0.375000",
                "-",
                "0.375000",
                "9.53",
                "6",
                "0",
                "3",
                "0",
                "3",
                "0",
            ]],
        ),
    )

    assert parsed.preview.invalid_rows == 1
    assert parsed.preview.valid_rows == 0
    assert {issue.code for issue in parsed.rows[0].errors} == {
        "invalid_datetime",
        "out_of_range",
    }
    assert parsed.rows[0].normalized["published_at"] is None


def test_required_identity_and_metric_fields_cannot_be_blank():
    daily = parse_source_file(
        "daily.csv",
        csv_bytes(DAILY_HEADERS, [["", "81"]]),
    )
    single = parse_source_file(
        "single.csv",
        csv_bytes(
            SINGLE_CONTENT_HEADERS,
            [["", "2026-07-07 13:36", "444", "0.1471", "0.6353", "2.8314"]],
        ),
    )

    assert {issue.code for issue in daily.rows[0].errors} == {"required_value_missing"}
    assert daily.rows[0].errors[0].field == "stat_date"
    assert {issue.code for issue in single.rows[0].errors} == {"required_value_missing"}
    assert single.rows[0].errors[0].field == "title"


def test_negative_counts_and_non_finite_numbers_are_invalid():
    daily = parse_source_file(
        "daily.csv",
        csv_bytes(DAILY_HEADERS, [["2026-07-18", "-1"]]),
    )
    single = parse_source_file(
        "single.csv",
        csv_bytes(
            SINGLE_CONTENT_HEADERS,
            [["作品 A", "2026-07-07 13:36", "444", "0.1471", "0.6353", "NaN"]],
        ),
    )

    assert {issue.code for issue in daily.rows[0].errors} == {"below_minimum"}
    assert {issue.code for issue in single.rows[0].errors} == {"invalid_number"}


def test_integer_overflow_and_oversized_titles_are_invalid():
    daily = parse_source_file(
        "daily.csv",
        csv_bytes(DAILY_HEADERS, [["2026-07-18", str(2**31)]]),
    )
    single = parse_source_file(
        "single.csv",
        csv_bytes(
            SINGLE_CONTENT_HEADERS,
            [["作" * 231, "2026-07-07 13:36", "444", "0.1471", "0.6353", "2.8314"]],
        ),
    )

    assert {issue.code for issue in daily.rows[0].errors} == {"integer_out_of_range"}
    assert {issue.code for issue in single.rows[0].errors} == {"value_too_long"}


def test_unknown_template_fails_clearly():
    with pytest.raises(ParseFailure, match="Unknown or unsupported template"):
        parse_source_file("unknown.csv", csv_bytes(["抖音号", "抖音名称"], [["a", "b"]]))


def test_duplicate_headers_fail_clearly():
    with pytest.raises(ParseFailure, match="Duplicate headers"):
        parse_source_file("dup.csv", csv_bytes(["日期", "日期"], [["2026-07-18", "81"]]))


def test_csv_short_row_is_rejected():
    payload = csv_bytes(["日期", "播放量"], [["2026-07-18"]])

    with pytest.raises(ParseFailure, match="expected 2 fields"):
        parse_source_file("daily.csv", payload)


def test_csv_long_row_is_rejected():
    payload = csv_bytes(["日期", "播放量"], [["2026-07-18", "81", "extra"]])

    with pytest.raises(ParseFailure, match="expected 2 fields"):
        parse_source_file("daily.csv", payload)


def test_xlsx_rows_with_unexpected_extra_cells_are_rejected():
    payload = workbook_bytes(
        DAILY_HEADERS,
        [["2026-07-18", "81", "unexpected"]],
    )

    with pytest.raises(ParseFailure, match="row 2 has 3 fields; expected at most 2"):
        parse_source_file("daily.xlsx", payload)


def test_workbooks_with_multiple_supported_worksheets_return_multiple_datasets():
    workbook = Workbook()
    first = workbook.active
    first.append(DAILY_HEADERS)
    first.append(["2026-07-18", 81])
    second = workbook.create_sheet("Another export")
    second.append(DAILY_HEADERS)
    second.append(["2026-07-19", 82])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    parsed = parse_source_file("multiple.xlsx", buffer.getvalue())

    assert [dataset.template_code for dataset in parsed.datasets] == [
        "douyin_daily_play_v1",
        "douyin_daily_play_v1",
    ]
    assert [dataset.rows[0].normalized["play"] for dataset in parsed.datasets] == [81, 82]


def test_header_only_file_preserves_template_code_in_preview():
    parsed = parse_source_file("works.csv", csv_bytes(WORK_LIST_HEADERS, []))

    assert parsed.template_code == "douyin_work_list_v1"
    assert parsed.preview.template_code == "douyin_work_list_v1"
    assert parsed.preview.total_rows == 0
    assert parsed.preview.valid_rows == 0
    assert parsed.preview.invalid_rows == 0


def test_formula_cells_are_rejected_before_openpyxl_load(monkeypatch: pytest.MonkeyPatch):
    payload = workbook_bytes(
        DAILY_HEADERS,
        [["2026-07-18", "81"]],
        formulas={(2, 2)},
    )
    monkeypatch.setattr(parser_module, "load_workbook", _fail_if_loaded)

    with pytest.raises(ParseFailure, match="Formula cells are not supported"):
        parse_source_file("daily.xlsx", payload)


def test_formula_scanner_does_not_treat_filter_tags_as_formulas():
    buffer = io.BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr(
            "xl/worksheets/sheet1.xml",
            b"<worksheet><filterColumn/><filters/></worksheet>",
        )

    with ZipFile(io.BytesIO(buffer.getvalue())) as archive:
        info = archive.getinfo("xl/worksheets/sheet1.xml")
        assert parser_module._entry_contains_formula(archive, info) is False


def test_external_links_are_rejected_before_openpyxl_load(monkeypatch: pytest.MonkeyPatch):
    payload = workbook_bytes(
        DAILY_HEADERS,
        [["2026-07-18", "81"]],
        extras={"xl/externalLinks/externalLink1.xml": b"<externalLink/>"},
    )
    monkeypatch.setattr(parser_module, "load_workbook", _fail_if_loaded)

    with pytest.raises(ParseFailure, match="external links"):
        parse_source_file("external.xlsx", payload)


def test_archive_entry_count_limit_rejects_before_decompression(
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(parser_module, "MAX_ARCHIVE_ENTRY_COUNT", 8)
    extras = {f"xl/junk/{index}.bin": b"x" for index in range(12)}
    payload = workbook_bytes(DAILY_HEADERS, [["2026-07-18", "81"]], extras=extras)
    monkeypatch.setattr(parser_module.ZipFile, "open", _fail_if_opened)
    monkeypatch.setattr(parser_module, "load_workbook", _fail_if_loaded)

    with pytest.raises(ParseFailure, match="too many archive entries"):
        parse_source_file("bomb.xlsx", payload)


def test_archive_total_uncompressed_limit_rejects_before_decompression(
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(parser_module, "MAX_ARCHIVE_TOTAL_UNCOMPRESSED_BYTES", 1_024)
    chunk = bytes(range(256)) * 3
    extras = {
        "xl/junk/a.bin": (chunk, ZIP_STORED),
        "xl/junk/b.bin": (chunk, ZIP_STORED),
        "xl/junk/c.bin": (chunk[:64], ZIP_STORED),
    }
    payload = workbook_bytes(DAILY_HEADERS, [["2026-07-18", "81"]], extras=extras)
    monkeypatch.setattr(parser_module.ZipFile, "open", _fail_if_opened)
    monkeypatch.setattr(parser_module, "load_workbook", _fail_if_loaded)

    with pytest.raises(ParseFailure, match="total uncompressed size"):
        parse_source_file("bomb.xlsx", payload)


def test_archive_entry_size_limit_rejects_before_decompression(
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(parser_module, "MAX_ARCHIVE_ENTRY_UNCOMPRESSED_BYTES", 512)
    extras = {"xl/junk/huge.bin": b"A" * 768}
    payload = workbook_bytes(DAILY_HEADERS, [["2026-07-18", "81"]], extras=extras)
    monkeypatch.setattr(parser_module.ZipFile, "open", _fail_if_opened)
    monkeypatch.setattr(parser_module, "load_workbook", _fail_if_loaded)

    with pytest.raises(ParseFailure, match="archive entry is too large"):
        parse_source_file("bomb.xlsx", payload)


def test_archive_compression_ratio_limit_rejects_before_decompression(
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(parser_module, "MAX_ARCHIVE_COMPRESSION_RATIO", 5)
    extras = {"xl/junk/high-ratio.bin": b"A" * 16_384}
    payload = workbook_bytes(DAILY_HEADERS, [["2026-07-18", "81"]], extras=extras)
    monkeypatch.setattr(parser_module.ZipFile, "open", _fail_if_opened)
    monkeypatch.setattr(parser_module, "load_workbook", _fail_if_loaded)

    with pytest.raises(ParseFailure, match="compression ratio"):
        parse_source_file("bomb.xlsx", payload)


def test_archive_entry_with_zero_compressed_size_is_rejected_before_decompression(
    monkeypatch: pytest.MonkeyPatch,
):
    class FakeZipInfo:
        filename = "xl/worksheets/sheet1.xml"
        file_size = 128
        compress_size = 0

        @staticmethod
        def is_dir() -> bool:
            return False

    class FakeZipFile:
        def __init__(self, *_args, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        @staticmethod
        def infolist():
            return [FakeZipInfo()]

        @staticmethod
        def open(*_args, **_kwargs):
            raise AssertionError("archive entry should not be decompressed")

    monkeypatch.setattr(parser_module, "ZipFile", FakeZipFile)
    monkeypatch.setattr(parser_module, "load_workbook", _fail_if_loaded)

    with pytest.raises(ParseFailure, match="compression ratio"):
        parse_source_file("bomb.xlsx", b"placeholder")


def test_duplicate_zip_entry_names_are_rejected_before_decompression(
    monkeypatch: pytest.MonkeyPatch,
):
    class FakeZipInfo:
        file_size = 32
        compress_size = 16

        def __init__(self, filename: str):
            self.filename = filename

        @staticmethod
        def is_dir() -> bool:
            return False

    class FakeZipFile:
        def __init__(self, *_args, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        @staticmethod
        def infolist():
            return [
                FakeZipInfo("[Content_Types].xml"),
                FakeZipInfo("xl/workbook.xml"),
                FakeZipInfo("xl/workbook.xml"),
            ]

        @staticmethod
        def open(*_args, **_kwargs):
            raise AssertionError("archive entry should not be decompressed")

    monkeypatch.setattr(parser_module, "ZipFile", FakeZipFile)
    monkeypatch.setattr(parser_module, "load_workbook", _fail_if_loaded)

    with pytest.raises(ParseFailure, match="duplicate archive entries"):
        parse_source_file("duplicate.xlsx", b"placeholder")


def test_malformed_csv_and_non_utf8_input_are_rejected():
    malformed = '"日期","播放量"\n"2026-07-18","81'.encode()
    with pytest.raises(ParseFailure, match="Malformed CSV"):
        parse_source_file("broken.csv", malformed)

    invalid_encoding = "日期,播放量\n2026-07-18,81".encode("utf-16le")
    with pytest.raises(ParseFailure, match="UTF-8"):
        parse_source_file("broken.csv", invalid_encoding)


def test_embedded_nul_and_unsupported_extensions_are_rejected():
    with pytest.raises(ParseFailure, match="Embedded NUL"):
        parse_source_file("daily.csv", b"\x00" + "日期,播放量\n2026-07-18,81".encode())

    with pytest.raises(ParseFailure, match="Unsupported file extension"):
        parse_source_file("daily.xls", b"not-an-xlsx")


def test_row_column_and_size_limits_are_enforced():
    too_many_rows = csv_bytes(
        DAILY_HEADERS,
        [[f"2026-07-{(day % 28) + 1:02d}", "1"] for day in range(10_001)],
    )
    with pytest.raises(ParseFailure, match="10,000"):
        parse_source_file("rows.csv", too_many_rows)

    too_many_columns = csv_bytes([f"列{i}" for i in range(101)], [[str(i) for i in range(101)]])
    with pytest.raises(ParseFailure, match="100 columns"):
        parse_source_file("columns.csv", too_many_columns)

    too_large = b"x" * (10 * 1024 * 1024 + 1)
    with pytest.raises(ParseFailure, match="10 MB"):
        parse_source_file("big.csv", too_large)


def test_file_adapter_builds_preview_contract():
    adapter = FileDataSourceAdapter()
    source = Path("works.csv")
    parsed = adapter.parse(
        {
            "filename": source.name,
            "data": csv_bytes(
                WORK_LIST_HEADERS,
                [[
                    "作品 A",
                    "2026-07-18 14:11:20",
                    "1min-视频",
                    "公开",
                    "81",
                    "0.0875",
                    "0.375",
                    "",
                    "0.375",
                    "9.53",
                    "6",
                    "0",
                    "3",
                    "0",
                    "3",
                    "0",
                ]],
            ),
        }
    )

    preview = adapter.preview(parsed.rows)

    assert adapter.detect({"filename": source.name, "data": b"test"}).matched is True
    assert preview.template_code == "douyin_work_list_v1"
    assert preview.total_rows == 1
    assert preview.valid_rows == 1
    assert preview.invalid_rows == 0


def test_file_adapter_preserves_template_code_for_header_only_preview():
    adapter = FileDataSourceAdapter()
    parsed = adapter.parse({"filename": "works.csv", "data": csv_bytes(WORK_LIST_HEADERS, [])})

    preview = adapter.preview(parsed.rows, template_code=parsed.template_code)

    assert preview.template_code == "douyin_work_list_v1"
    assert preview.total_rows == 0
